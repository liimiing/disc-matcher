#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Discogs 音乐专辑匹配器
自动从Discogs搜索专辑信息，下载封面和元数据，并整理文件夹

使用前请配置Discogs Token:
1. 访问 https://www.discogs.com/settings/developers
2. 生成Personal Access Token
3. 在DiscMatcherApp类中找到DISCOGS_TOKEN变量，将Token填入
"""

import os
import sys
import json
import time
import requests
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import threading
from urllib.parse import quote
from PIL import Image, ImageTk
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import subprocess
import shlex
import locale
from translations import TRANSLATIONS


class LanguageManager:
    """多语言管理器"""
    
    def __init__(self):
        self.current_lang = 'auto'
        self.translations = TRANSLATIONS
    
    def detect_system_language(self):
        """检测系统语言"""
        try:
            lang, _ = locale.getdefaultlocale()
            if lang:
                lang_lower = lang.lower()
                if lang_lower.startswith('zh_tw') or lang_lower.startswith('zh_hant'):
                    return 'zh_TW'
                elif lang_lower.startswith('zh'):
                    return 'zh_CN'
                elif lang_lower.startswith('es'):
                    return 'es_ES'
                elif lang_lower.startswith('pt'):
                    return 'pt_BR'
                elif lang_lower.startswith('fr'):
                    return 'fr_FR'
                elif lang_lower.startswith('de'):
                    return 'de_DE'
                elif lang_lower.startswith('ja'):
                    return 'ja_JP'
                elif lang_lower.startswith('ko'):
                    return 'ko_KR'
                elif lang_lower.startswith('ru'):
                    return 'ru_RU'
                elif lang_lower.startswith('ar'):
                    return 'ar_SA'
                elif lang_lower.startswith('hi'):
                    return 'hi_IN'
                else:
                    return 'en_US'
        except:
            pass
        return 'en_US'  # 默认英文
    
    def set_language(self, lang_code):
        """设置语言"""
        if lang_code == 'auto':
            self.current_lang = self.detect_system_language()
        else:
            self.current_lang = lang_code
    
    def get_current_lang(self):
        """获取当前语言代码"""
        if self.current_lang == 'auto':
            return self.detect_system_language()
        return self.current_lang
    
    def t(self, key, **kwargs):
        """翻译文本"""
        lang = self.get_current_lang()
        text = self.translations.get(lang, self.translations['en_US']).get(key, key)
        # 支持格式化字符串
        if kwargs:
            try:
                text = text.format(**kwargs)
            except:
                pass
        return text
    
    def get_available_languages(self):
        """获取可用语言列表"""
        return [
            ('auto', self.t('auto')),
            ('zh_CN', '简体中文'),
            ('zh_TW', '繁體中文'),
            ('en_US', 'English'),
            ('es_ES', 'Español'),
            ('pt_BR', 'Português'),
            ('fr_FR', 'Français'),
            ('de_DE', 'Deutsch'),
            ('ja_JP', '日本語'),
            ('ko_KR', '한국어'),
            ('ru_RU', 'Русский'),
            ('ar_SA', 'العربية'),
            ('hi_IN', 'हिन्दी'),
        ]


class DiscogsAPI:
    """Discogs API 客户端"""
    
    BASE_URL = "https://api.discogs.com"
    
    def __init__(self, token: str):
        self.token = token
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'DiscMatcher/1.0',
            'Authorization': f'Discogs token={token}'
        })
    
    def search(self, query: str) -> List[Dict]:
        """搜索专辑"""
        # 清理查询字符串
        cleaned_query = query.replace('[', '').replace(']', '').replace('(', '').replace(')', '')
        cleaned_query = cleaned_query.replace('.', ' ').replace('_', ' ').replace('-', ' ')
        cleaned_query = ' '.join(cleaned_query.split())
        
        url = f"{self.BASE_URL}/database/search"
        params = {
            'q': cleaned_query,
            'type': 'release',
            'token': self.token
        }
        
        try:
            response = self.session.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            return data.get('results', [])
        except requests.exceptions.RequestException as e:
            print(f"搜索错误: {e}")
            return []
    
    def get_release_details(self, release_id: int) -> Optional[Dict]:
        """获取专辑详细信息"""
        url = f"{self.BASE_URL}/releases/{release_id}"
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"获取详情错误: {e}")
            return None
    
    def download_image(self, url: str, save_path: Path) -> bool:
        """下载图片"""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            img = Image.open(io.BytesIO(response.content))
            img.save(save_path)
            return True
        except Exception as e:
            print(f"下载图片错误: {e}")
            return False


class AlbumInfo:
    """专辑信息类"""
    
    def __init__(self, release_data: Dict):
        self.release_id = release_data.get('id')
        self.title = release_data.get('title', '')
        self.year = release_data.get('year', '')
        self.cover_image = release_data.get('cover_image', '')
        self.thumb = release_data.get('thumb', '')
        
        # 解析艺术家和专辑名
        if ' - ' in self.title:
            parts = self.title.split(' - ', 1)
            self.artist = parts[0].strip()
            self.album_name = parts[1].strip()
        else:
            self.artist = ''
            self.album_name = self.title
        
        # 标签信息
        self.labels = release_data.get('label', [])
        if isinstance(self.labels, list):
            self.label_names = [l.get('name', '') if isinstance(l, dict) else str(l) for l in self.labels]
        else:
            self.label_names = [str(self.labels)] if self.labels else []
        
        self.catalog_no = release_data.get('catno', '')
        self.country = release_data.get('country', '')
        self.genre = release_data.get('genre', [])
        self.style = release_data.get('style', [])
        self.format = release_data.get('format', [])
        
        # 详细信息（需要额外请求）
        self.details = None
        self.notes = ''
        self.tracklist = []  # 曲目表
        self.images = []  # 所有图片URL
    
    def sanitize_filename(self, filename: str) -> str:
        """清理Windows文件名中的非法字符"""
        # Windows不允许的字符: < > : " / \ | ? *
        illegal_chars = r'[<>:"/\\|?*]'
        # 替换为下划线
        sanitized = re.sub(illegal_chars, '_', filename)
        # 移除首尾空格和点号
        sanitized = sanitized.strip(' .')
        # 移除连续的下划线
        sanitized = re.sub(r'_+', '_', sanitized)
        return sanitized
    
    def get_suggested_folder_name(self) -> str:
        """生成建议的文件夹名：音乐人 -年份- 专辑名"""
        parts = []
        if self.artist:
            parts.append(self.sanitize_filename(self.artist))
        if self.year:
            parts.append(str(self.year))
        if self.album_name:
            parts.append(self.sanitize_filename(self.album_name))
        
        if parts:
            suggested = ' - '.join(parts)
            # 再次清理整个字符串
            return self.sanitize_filename(suggested)
        return self.sanitize_filename(self.title)
    
    def to_dict(self) -> Dict:
        """转换为字典用于Excel导出和JSON保存"""
        return {
            '音乐人': self.artist,
            '专辑名': self.album_name,
            '出版年份': self.year,
            '唱片厂牌': ', '.join(self.label_names),
            '厂牌编号': self.catalog_no,
            '音乐风格': ', '.join(self.genre) if isinstance(self.genre, list) else str(self.genre),
            '风格标签': ', '.join(self.style) if isinstance(self.style, list) else str(self.style),
            '备注信息': self.notes,
            'Discogs ID': self.release_id,
            '国家': self.country,
            '曲目表': self.tracklist
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'AlbumInfo':
        """从字典创建AlbumInfo对象（用于从JSON加载）"""
        # 处理标题，确保格式正确
        artist = data.get('音乐人', '').strip()
        album_name = data.get('专辑名', '').strip()
        if artist and album_name:
            title = f"{artist} - {album_name}"
        elif album_name:
            title = album_name
        elif artist:
            title = artist
        else:
            title = ''
        
        # 处理音乐风格和风格标签
        genre_str = data.get('音乐风格', '')
        if isinstance(genre_str, list):
            genre = genre_str
        elif genre_str:
            genre = [g.strip() for g in genre_str.split(',') if g.strip()]
        else:
            genre = []
        
        style_str = data.get('风格标签', '')
        if isinstance(style_str, list):
            style = style_str
        elif style_str:
            style = [s.strip() for s in style_str.split(',') if s.strip()]
        else:
            style = []
        
        # 处理唱片厂牌
        label_str = data.get('唱片厂牌', '')
        if isinstance(label_str, list):
            label = [{'name': str(l)} for l in label_str]
        elif label_str:
            label = [{'name': name.strip()} for name in label_str.split(',') if name.strip()]
        else:
            label = []
        
        # 创建一个临时的release_data结构
        release_data = {
            'id': data.get('Discogs ID'),
            'title': title,
            'year': str(data.get('出版年份', '')) if data.get('出版年份') else '',
            'catno': data.get('厂牌编号', ''),
            'country': data.get('国家', ''),
            'genre': genre,
            'style': style,
            'label': label,
            'cover_image': '',  # JSON中不保存图片URL，因为图片已下载
            'thumb': ''
        }
        
        album_info = cls(release_data)
        album_info.notes = data.get('备注信息', '')
        
        # 处理曲目表
        tracklist_data = data.get('曲目表', [])
        if isinstance(tracklist_data, list):
            album_info.tracklist = tracklist_data
        else:
            album_info.tracklist = []
        
        return album_info


class ModernMenu:
    """
    一个用于替代 tk.Menu 的自定义扁平化菜单类。
    完全支持深色模式，无强制白边。
    """
    def __init__(self, root, bg, fg, active_bg, active_fg):
        self.root = root
        self.bg = bg
        self.fg = fg
        self.active_bg = active_bg
        self.active_fg = active_fg
        
        self.items = []  # 存储菜单项数据
        self.menu_window = None  # 实际显示的窗口引用
        self.shadow_window = None  # 阴影窗口引用

    def add_command(self, label, command):
        """添加菜单项"""
        self.items.append({'type': 'command', 'label': label, 'command': command})

    def add_separator(self):
        """添加分隔线"""
        self.items.append({'type': 'separator'})

    def post(self, x, y):
        """在指定位置显示菜单"""
        self._destroy_menu()  # 如果已存在则销毁
        
        # 计算边框颜色（比背景色稍亮）
        def lighten_color(color, factor=0.15):
            """将颜色变亮"""
            if isinstance(color, str) and color.startswith('#'):
                r = int(color[1:3], 16)
                g = int(color[3:5], 16)
                b = int(color[5:7], 16)
                r = min(255, int(r + (255 - r) * factor))
                g = min(255, int(g + (255 - g) * factor))
                b = min(255, int(b + (255 - b) * factor))
                return f"#{r:02x}{g:02x}{b:02x}"
            return color
        
        border_color = lighten_color(self.bg, 0.2)  # 边框颜色稍亮
        shadow_color = "#000000"  # 阴影颜色（黑色半透明）
        shadow_offset = 3  # 阴影偏移量
        
        # 创建主菜单窗口（先创建，以便计算尺寸）
        self.menu_window = tk.Toplevel(self.root)
        self.menu_window.overrideredirect(True)  # 移除系统标题栏和边框
        self.menu_window.attributes('-topmost', True)  # 保持在最前
        self.menu_window.config(bg=border_color)
        
        # 边框层（使用稍亮的颜色）
        border_frame = tk.Frame(self.menu_window, bg=border_color, bd=1)
        border_frame.pack(fill='both', expand=True, padx=1, pady=1)
        
        # 内部容器（实际内容区域）
        content_frame = tk.Frame(border_frame, bg=self.bg)
        content_frame.pack(fill='both', expand=True)

        # 构建菜单项
        for item in self.items:
            if item['type'] == 'command':
                btn = tk.Label(content_frame, 
                               text=f"  {item['label']}  ", 
                               bg=self.bg, 
                               fg=self.fg,
                               font=('Arial', 9),
                               anchor='w',
                               pady=6,
                               padx=10)
                btn.pack(fill='x', expand=True)
                
                # 绑定事件
                # 使用闭包捕获当前的 cmd 和 btn，避免闭包陷阱
                cmd = item['command']  # 在循环中先保存命令
                def make_handlers(button, command):
                    def on_enter(e): 
                        button.config(bg=self.active_bg, fg=self.active_fg)
                    def on_leave(e): 
                        button.config(bg=self.bg, fg=self.fg)
                    def on_click(e):
                        self._destroy_menu()
                        if command: 
                            command()
                    return on_enter, on_leave, on_click
                
                on_enter, on_leave, on_click = make_handlers(btn, cmd)
                btn.bind('<Enter>', on_enter)
                btn.bind('<Leave>', on_leave)
                btn.bind('<Button-1>', on_click)
                
            elif item['type'] == 'separator':
                sep = tk.Frame(content_frame, bg="#404040", height=1)
                sep.pack(fill='x', pady=4, padx=5)

        # 更新窗口以确保尺寸计算正确
        self.menu_window.update_idletasks()
        
        # 获取菜单实际尺寸
        menu_width = self.menu_window.winfo_reqwidth()
        menu_height = self.menu_window.winfo_reqheight()
        
        # 创建阴影窗口（投影效果）- 在菜单窗口之后创建，确保在下方
        shadow_width = menu_width + shadow_offset
        shadow_height = menu_height + shadow_offset
        self.shadow_window = tk.Toplevel(self.root)
        self.shadow_window.overrideredirect(True)
        self.shadow_window.attributes('-topmost', False)  # 阴影在菜单下方
        self.shadow_window.attributes('-alpha', 0.25)  # 半透明阴影
        self.shadow_window.config(bg=shadow_color)
        self.shadow_window.geometry(f"{shadow_width}x{shadow_height}+{x+shadow_offset}+{y+shadow_offset}")
        
        # 确保菜单窗口在最上层
        self.menu_window.attributes('-topmost', True)
        
        # 设置主菜单窗口位置
        self.menu_window.geometry(f"{menu_width}x{menu_height}+{x}+{y}")
        
        # 绑定主窗口事件，当窗口移动、失去焦点等时关闭菜单
        self._bind_root_events()
        
        # 点击菜单外部关闭菜单的逻辑
        # 绑定全局点击事件，如果点击不在菜单内，则关闭
        # 注意：这需要延时绑定，否则触发菜单的那次点击会立即关闭菜单
        self.menu_window.after(100, lambda: self.root.bind_all("<Button-1>", self._check_click_outside))

    def _check_click_outside(self, event):
        """检查点击是否在菜单外部"""
        if self.menu_window:
            try:
                # 获取菜单窗口和阴影窗口的坐标范围
                menu_x = self.menu_window.winfo_x()
                menu_y = self.menu_window.winfo_y()
                menu_w = self.menu_window.winfo_width()
                menu_h = self.menu_window.winfo_height()
                
                shadow_x = self.shadow_window.winfo_x() if self.shadow_window else menu_x
                shadow_y = self.shadow_window.winfo_y() if self.shadow_window else menu_y
                shadow_w = self.shadow_window.winfo_width() if self.shadow_window else menu_w
                shadow_h = self.shadow_window.winfo_height() if self.shadow_window else menu_h
                
                # 如果点击坐标不在菜单或阴影范围内，销毁菜单
                in_menu = (menu_x <= event.x_root <= menu_x + menu_w and 
                          menu_y <= event.y_root <= menu_y + menu_h)
                in_shadow = (shadow_x <= event.x_root <= shadow_x + shadow_w and 
                            shadow_y <= event.y_root <= shadow_y + shadow_h)
                
                if not (in_menu or in_shadow):
                    self._destroy_menu()
            except:
                self._destroy_menu()

    def _bind_root_events(self):
        """绑定主窗口事件，当窗口移动或失去焦点时关闭菜单"""
        # 保存窗口初始位置，用于检测移动
        self._last_root_x = self.root.winfo_x()
        self._last_root_y = self.root.winfo_y()
        
        # 窗口配置改变事件（移动、大小改变）
        self._configure_handler = self.root.bind('<Configure>', self._on_root_configure)
        # 窗口失去焦点事件
        self._focus_out_handler = self.root.bind('<FocusOut>', self._on_root_focus_out)
        # 窗口最小化或隐藏事件
        self._unmap_handler = self.root.bind('<Unmap>', self._on_root_unmap)
    
    def _unbind_root_events(self):
        """解除主窗口事件绑定"""
        try:
            if hasattr(self, '_configure_handler'):
                self.root.unbind('<Configure>', self._configure_handler)
            if hasattr(self, '_focus_out_handler'):
                self.root.unbind('<FocusOut>', self._focus_out_handler)
            if hasattr(self, '_unmap_handler'):
                self.root.unbind('<Unmap>', self._unmap_handler)
        except:
            pass
    
    def _on_root_configure(self, event):
        """主窗口配置改变（移动、大小改变）时关闭菜单"""
        # 只处理主窗口本身的事件，忽略子控件的事件
        if event.widget == self.root:
            try:
                current_x = self.root.winfo_x()
                current_y = self.root.winfo_y()
                # 检测窗口是否真的移动了（避免子控件变化导致的误触发）
                if (current_x != self._last_root_x or current_y != self._last_root_y):
                    self._last_root_x = current_x
                    self._last_root_y = current_y
                    self._destroy_menu()
            except:
                pass
    
    def _on_root_focus_out(self, event):
        """主窗口失去焦点时关闭菜单"""
        # 检查是否真的失去了焦点（切换到其他应用）
        try:
            # 延迟检查，避免菜单点击时误触发
            self.root.after(50, self._check_focus_lost)
        except:
            self._destroy_menu()
    
    def _check_focus_lost(self):
        """检查窗口是否真的失去了焦点"""
        try:
            # 如果菜单还存在，且窗口不在前台，则关闭菜单
            if self.menu_window:
                focused = self.root.focus_displayof()
                if focused is None or str(focused) != str(self.root):
                    self._destroy_menu()
        except:
            pass
    
    def _on_root_unmap(self, event):
        """主窗口最小化或隐藏时关闭菜单"""
        if event.widget == self.root:
            self._destroy_menu()
    
    def _destroy_menu(self):
        """销毁菜单窗口"""
        if self.menu_window:
            self.root.unbind_all("<Button-1>")  # 解除全局绑定
            self._unbind_root_events()  # 解除主窗口事件绑定
            self.menu_window.destroy()
            self.menu_window = None
        if hasattr(self, 'shadow_window') and self.shadow_window:
            try:
                self.shadow_window.destroy()
            except:
                pass
            self.shadow_window = None


class DiscMatcherApp:
    """主应用程序"""
    
    def __init__(self, root):
        self.root = root
        # 标题将在setup_ui中根据语言设置
        self.root.geometry("1200x800")
        
        self.discogs_api = None
        self.root_folder = None
        self.album_folders = []  # List of (folder_path, folder_name, album_info)
        self.processing_thread = None
        self.waiting_for_selection = threading.Event()  # 用于等待用户选择
        self.selection_result = None  # 存储用户选择的结果
        self.selection_dialog_active = False  # 标记选择对话框是否正在显示
        self.open_dialogs = []  # 保存所有打开的对话框引用，用于跟随主窗体移动
        self.open_toasts = []  # 保存所有打开的toast引用，用于跟随主窗体移动
        self._last_root_position = None  # 记录主窗体上次位置，用于检测移动
        self._position_check_job = None  # 位置检查定时器任务ID
        
        # Discogs Token - 从配置文件读取
        self.DISCOGS_TOKEN = None
        
        # 配置文件路径（保存在程序目录）
        self.config_file = Path(__file__).parent / 'config.json'
        
        # 初始化语言管理器
        self.lang = LanguageManager()
        
        # 先加载配置（包括语言设置和Discogs Token）
        self.load_config()
        
        # 初始化Discogs API（在加载配置后）
        if self.DISCOGS_TOKEN and self.DISCOGS_TOKEN != "YOUR_DISCOGS_TOKEN_HERE":
            self.discogs_api = DiscogsAPI(self.DISCOGS_TOKEN)
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置用户界面"""
        # 深空灰色主题配色
        self.bg_color = "#1E1E1E"  # 主背景色（深空灰）
        self.secondary_bg = "#2D2D2D"  # 次要背景色
        self.accent_bg = "#3C3C3C"  # 强调背景色
        self.text_color = "#E0E0E0"  # 文本颜色
        self.accent_color = "#4A9EFF"  # 强调色（蓝色）
        self.success_color = "#4CAF50"  # 成功色（绿色）
        self.error_color = "#F44336"  # 错误色（红色）
        
        # 设置根窗口背景
        self.root.configure(bg=self.bg_color)
        
        # 设置窗口标题（根据当前语言）
        self.root.title(self.lang.t('app_title') + ' V3.4')
        
        # 配置ttk样式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配置Frame样式
        style.configure('TFrame', background=self.bg_color)
        style.configure('Toolbar.TFrame', background=self.secondary_bg)
        
        # 配置Label样式
        style.configure('TLabel', background=self.bg_color, foreground=self.text_color)
        style.configure('Status.TLabel', background=self.secondary_bg, foreground=self.text_color)
        
        # 配置Button样式
        style.configure('TButton', 
                        background=self.accent_bg,
                        foreground=self.text_color,
                        borderwidth=0,
                        focuscolor='none',
                        bordercolor=self.secondary_bg,
                        lightcolor=self.accent_bg,
                        darkcolor=self.accent_bg)
        style.map('TButton',
                  background=[('active', self.accent_color), ('pressed', '#3A8EEF')],
                  foreground=[('active', 'white')],
                  bordercolor=[('active', self.secondary_bg), ('pressed', self.secondary_bg)])
        
        # 配置Progressbar样式
        style.configure('TProgressbar',
                        background=self.accent_color,
                        troughcolor=self.secondary_bg,
                        borderwidth=0,
                        lightcolor=self.accent_color,
                        darkcolor=self.accent_color)
        
        # 配置Treeview样式 - 移除白色边框
        # 修改布局，移除Treeview.field的边框元素
        try:
            # 获取默认布局
            default_layout = style.layout('Treeview')
            # 创建新布局，移除field的边框
            new_layout = []
            for element in default_layout:
                if isinstance(element, tuple) and len(element) == 2:
                    element_name, element_options = element
                    if element_name == 'Treeview.field':
                        # 移除field的边框，设置border为0
                        new_options = dict(element_options)
                        new_options['border'] = '0'
                        new_layout.append((element_name, new_options))
                    else:
                        new_layout.append(element)
                else:
                    new_layout.append(element)
            style.layout('Treeview', new_layout)
        except:
            # 如果获取默认布局失败，使用简化布局（无边框）
            treeview_layout = [
                ('Treeview.field', {'sticky': 'nswe', 'border': '0'}),
                ('Treeview.padding', {'sticky': 'nswe', 'children': [
                    ('Treeitem', {'sticky': 'nswe'})
                ]})
            ]
            style.layout('Treeview', treeview_layout)
        
        style.configure('Treeview',
                       background=self.secondary_bg,
                       foreground=self.text_color,
                       fieldbackground=self.secondary_bg,
                       borderwidth=0,
                       bordercolor=self.secondary_bg)
        style.configure('Treeview.Heading',
                       background=self.accent_bg,
                       foreground=self.text_color,
                       relief='flat',
                       borderwidth=0,
                       bordercolor=self.secondary_bg)
        style.map('Treeview',
                  background=[('selected', self.accent_color)],
                  foreground=[('selected', 'white')])
        
        # 配置自定义Treeview样式（用于移除白色边框）
        # 修改布局，移除Treeview.field的边框元素
        try:
            default_layout = style.layout('Treeview')
            # 创建新布局，移除field的边框
            new_layout = []
            for element in default_layout:
                if element[0] == 'Treeview.field':
                    # 移除field的边框，只保留内容
                    new_layout.append(('Treeview.field', {'sticky': 'nswe', 'border': '0'}))
                else:
                    new_layout.append(element)
            style.layout('Custom.Treeview', new_layout)
        except:
            # 如果获取默认布局失败，使用简化布局
            treeview_layout = [
                ('Treeview.field', {'sticky': 'nswe', 'border': '0'}),
                ('Treeview.padding', {'sticky': 'nswe', 'children': [
                    ('Treeitem', {'sticky': 'nswe'})
                ]})
            ]
            style.layout('Custom.Treeview', treeview_layout)
        
        style.configure('Custom.Treeview',
                       background=self.secondary_bg,
                       foreground=self.text_color,
                       fieldbackground=self.secondary_bg,
                       borderwidth=0,
                       bordercolor=self.secondary_bg)
        style.configure('Custom.Treeview.Heading',
                       background=self.accent_bg,
                       foreground=self.text_color,
                       relief='flat',
                       borderwidth=0,
                       bordercolor=self.secondary_bg)
        style.map('Custom.Treeview',
                  background=[('selected', self.accent_color)],
                  foreground=[('selected', 'white')])
        
        # 配置Scrollbar样式
        style.configure('TScrollbar',
                        background=self.accent_bg,
                        troughcolor=self.secondary_bg,
                        borderwidth=0,
                        arrowcolor=self.text_color,
                        bordercolor=self.secondary_bg,
                        lightcolor=self.accent_bg,
                        darkcolor=self.accent_bg)
        
        # 顶部工具栏
        toolbar = ttk.Frame(self.root, padding="10", style='Toolbar.TFrame')
        toolbar.pack(fill=tk.X)
        
        # 显示Token状态
        token_status = self.lang.t('token_configured') if (self.DISCOGS_TOKEN and self.DISCOGS_TOKEN != "YOUR_DISCOGS_TOKEN_HERE") else self.lang.t('token_not_configured')
        token_color = self.success_color if token_status == self.lang.t('token_configured') else self.error_color
        self.token_label_ref = ttk.Label(toolbar, text=f"{self.lang.t('token_status')}: {token_status}", foreground=token_color, background=self.secondary_bg)
        self.token_label_ref.pack(side=tk.LEFT, padx=5)
        
        self.select_btn_ref = ttk.Button(toolbar, text=self.lang.t('select_folder'), command=self.select_folder)
        self.select_btn_ref.pack(side=tk.LEFT, padx=5)
        self.start_btn_ref = ttk.Button(toolbar, text=self.lang.t('start_processing'), command=self.start_processing)
        self.start_btn_ref.pack(side=tk.LEFT, padx=5)
        self.batch_rename_btn_ref = ttk.Button(toolbar, text=self.lang.t('batch_rename'), command=self.batch_rename)
        self.batch_rename_btn_ref.pack(side=tk.LEFT, padx=5)
        self.export_btn_ref = ttk.Button(toolbar, text=self.lang.t('export_excel'), command=self.export_excel)
        self.export_btn_ref.pack(side=tk.LEFT, padx=5)
        
        # 语言选择下拉框
        self.lang_label = ttk.Label(toolbar, text=f"{self.lang.t('language')}:", background=self.secondary_bg)
        self.lang_label.pack(side=tk.LEFT, padx=(20, 5))
        
        # 保存语言选项和代码的映射
        lang_options = []
        self.lang_values = []
        for code, name in self.lang.get_available_languages():
            lang_options.append(name)
            self.lang_values.append(code)
        
        # 找到当前语言对应的显示名称
        current_index = 0
        if self.lang.current_lang in self.lang_values:
            current_index = self.lang_values.index(self.lang.current_lang)
        current_display_name = lang_options[current_index]
        
        self.lang_var = tk.StringVar(value=current_display_name)
        self.lang_combo = ttk.Combobox(toolbar, textvariable=self.lang_var, width=12, state='readonly')
        self.lang_combo['values'] = lang_options
        self.lang_combo.current(current_index)
        self.lang_combo.pack(side=tk.LEFT, padx=5)
        self.lang_combo.bind('<<ComboboxSelected>>', self.on_language_changed)
        
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(toolbar, variable=self.progress_var, maximum=100, length=200)
        self.progress_bar.pack(side=tk.LEFT, padx=10)
        
        # 状态栏
        self.status_var = tk.StringVar(value=self.lang.t('ready'))
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='flat', style='Status.TLabel')
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 主内容区域
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建Treeview显示列表
        # 使用固定的列标识符（英文），表头显示时使用翻译文本
        self.tree_columns = ('folder_name', 'artist', 'album_name', 'year', 'status', 'suggested_name')
        
        self.tree = ttk.Treeview(main_frame, columns=self.tree_columns, show='headings', height=20, style='Custom.Treeview')
        
        # 配置不同状态的tag样式（深灰色系，略有区别）
        # 待处理 - 默认深灰色
        self.tree.tag_configure('pending', background=self.secondary_bg)
        # 搜索中 - 稍亮的深灰色
        self.tree.tag_configure('searching', background='#353535')
        # 已完成 - 稍亮的深灰色（与搜索中略有不同）
        self.tree.tag_configure('completed', background='#333333')
        # 未找到 - 稍暗的深灰色
        self.tree.tag_configure('notfound', background='#2A2A2A')
        
        # 更新表头文本（使用翻译）
        self.update_tree_headers()
        
        self.tree.column('folder_name', width=200)
        self.tree.column('suggested_name', width=250)
        
        # 设置年份和状态列居中显示
        self.tree.column('year', anchor='center', width=80)
        self.tree.column('status', anchor='center', width=80)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 移除Treeview的白色高亮边框（通过底层widget配置）
        try:
            # 获取Treeview的底层widget并移除高亮
            for widget in self.tree.winfo_children():
                if isinstance(widget, tk.Widget):
                    widget.configure(highlightthickness=0, highlightbackground=self.secondary_bg)
        except:
            pass
        
        # 右键菜单 - 使用自定义ModernMenu类以彻底移除白边
        self.context_menu = ModernMenu(self.root, 
                                       bg=self.secondary_bg, 
                                       fg=self.text_color,
                                       active_bg=self.accent_color,
                                       active_fg='white')
        
        # 更新右键菜单文本（使用翻译）
        self.update_context_menu()
        
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.on_double_click)
        
        # 如果配置中已有文件夹路径，自动扫描
        if self.root_folder and self.root_folder.exists():
            self.scan_folders()
        
        # 启动位置检查定时器，使对话框和toast跟随主窗体移动
        self._start_position_tracker()
    
    def _start_position_tracker(self):
        """启动位置跟踪器，定期检查主窗体位置变化"""
        try:
            # 记录初始位置
            self._last_root_position = (self.root.winfo_x(), self.root.winfo_y())
            # 启动定时检查
            self._check_position_change()
        except:
            pass
    
    def _check_position_change(self):
        """检查主窗体位置是否改变，如果改变则更新所有对话框和toast位置"""
        try:
            # 检查主窗体是否还存在
            if not self.root.winfo_exists():
                return
            
            current_x = self.root.winfo_x()
            current_y = self.root.winfo_y()
            current_pos = (current_x, current_y)
            
            # 如果位置改变了
            if self._last_root_position is None or self._last_root_position != current_pos:
                self._last_root_position = current_pos
                # 更新所有对话框位置
                self._update_all_dialogs()
                # 更新所有toast位置
                self._update_all_toasts()
            
            # 100ms后再次检查（10次/秒，足够流畅）
            self._position_check_job = self.root.after(100, self._check_position_change)
        except tk.TclError:
            # 主窗体可能已关闭，停止跟踪
            return
        except Exception:
            # 如果出错，尝试重新启动
            try:
                if self.root.winfo_exists():
                    self._position_check_job = self.root.after(100, self._check_position_change)
            except:
                pass
    
    def center_dialog(self, dialog, dialog_width, dialog_height):
        """将对话框相对于主窗体居中显示"""
        try:
            dialog.update_idletasks()
            # 获取主窗体位置和大小
            root_x = self.root.winfo_x()
            root_y = self.root.winfo_y()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            
            # 计算对话框位置（相对于主窗体居中）
            x = root_x + (root_width // 2) - (dialog_width // 2)
            y = root_y + (root_height // 2) - (dialog_height // 2)
            
            dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        except:
            pass
    
    def center_toast(self, toast, toast_width, toast_height):
        """将toast相对于主窗体居中显示"""
        try:
            toast.update_idletasks()
            # 获取主窗体位置和大小
            root_x = self.root.winfo_x()
            root_y = self.root.winfo_y()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            
            # 计算toast位置（相对于主窗体居中，稍微偏上）
            x = root_x + (root_width // 2) - (toast_width // 2)
            y = root_y + (root_height // 3) - (toast_height // 2)  # 在主窗体上方1/3处
            
            toast.geometry(f"{toast_width}x{toast_height}+{x}+{y}")
        except:
            pass
    
    def _update_all_dialogs(self):
        """更新所有打开的对话框位置"""
        for dialog_info in self.open_dialogs[:]:  # 使用切片复制，避免迭代时修改列表
            try:
                dialog, width, height = dialog_info
                if dialog.winfo_exists():
                    self.center_dialog(dialog, width, height)
                else:
                    # 对话框已关闭，从列表中移除
                    self.open_dialogs.remove(dialog_info)
            except tk.TclError:
                # 对话框可能已销毁，从列表中移除
                try:
                    self.open_dialogs.remove(dialog_info)
                except:
                    pass
            except Exception:
                pass
    
    def _update_all_toasts(self):
        """更新所有打开的toast位置"""
        for toast_info in self.open_toasts[:]:  # 使用切片复制，避免迭代时修改列表
            try:
                toast, width, height = toast_info
                if toast.winfo_exists():
                    self.center_toast(toast, width, height)
                else:
                    # toast已关闭，从列表中移除
                    self.open_toasts.remove(toast_info)
            except tk.TclError:
                # toast可能已销毁，从列表中移除
                try:
                    self.open_toasts.remove(toast_info)
                except:
                    pass
            except Exception:
                pass
    
    def show_toast(self, message: str, duration: int = 2000):
        """显示渐隐渐现的提示消息（支持多语言，跟随主窗体移动）"""
        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)  # 移除窗口装饰
        toast.transient(self.root)  # 设置为子窗口
        toast.attributes('-topmost', True)  # 确保在最上层显示
        
        # 深蓝色背景
        toast_bg = "#1E3A5F"  # 深蓝色
        toast.configure(bg=toast_bg)
        
        # 创建临时标签来测量文本宽度
        temp_label = tk.Label(toast, 
                             text=message,
                             font=('Arial', 11),
                             padx=20,
                             pady=20,
                             bg=toast_bg)
        temp_label.pack()
        toast.update_idletasks()
        
        # 获取文本实际需要的宽度
        text_width = temp_label.winfo_reqwidth()
        temp_label.destroy()
        
        # 设置宽度限制（最小200，最大600，根据内容自适应）
        min_width = 200
        max_width = 600
        width = max(min_width, min(text_width, max_width))
        height = 80
        
        # 先设置大小
        toast.geometry(f"{width}x{height}")
        toast.update_idletasks()
        
        # 相对于主窗体居中显示
        self.center_toast(toast, width, height)
        
        # 保存toast引用，以便跟随主窗体移动
        self.open_toasts.append((toast, width, height))
        
        # 创建标签（使用深蓝色背景）
        label = tk.Label(toast, 
                        text=message,
                        bg=toast_bg,
                        fg='white',  # 白色文字在深蓝背景上更清晰
                        font=('Arial', 11),
                        padx=20,
                        pady=20,
                        wraplength=width-40)  # 设置自动换行宽度
        label.pack(fill=tk.BOTH, expand=True)
        
        toast.update_idletasks()
        
        # 初始透明度设置为0，然后渐显
        toast.attributes('-alpha', 0.0)
        toast.deiconify()  # 确保窗口显示
        
        # 渐显动画
        def fade_in(step=0):
            try:
                if not toast.winfo_exists():
                    return
                if step <= 10:
                    alpha = step / 10.0
                    toast.attributes('-alpha', alpha)
                    toast.after(20, lambda: fade_in(step + 1))
                else:
                    # 等待指定时间后开始渐隐
                    toast.after(duration, fade_out)
            except:
                pass
        
        # 渐隐动画
        def fade_out(step=0):
            try:
                if not toast.winfo_exists():
                    return
                if step <= 10:
                    alpha = 1.0 - (step / 10.0)
                    toast.attributes('-alpha', alpha)
                    toast.after(20, lambda: fade_out(step + 1))
                else:
                    # 从列表中移除
                    try:
                        self.open_toasts.remove((toast, width, height))
                    except:
                        pass
                    try:
                        toast.destroy()
                    except:
                        pass
            except:
                pass
        
        # 延迟一小段时间后开始渐显，确保窗口完全创建
        toast.after(10, fade_in)
    
    def load_config(self):
        """加载配置文件，自动打开上次使用的文件夹和语言设置"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # 加载语言设置
                    lang_setting = config.get('language', 'auto')
                    self.lang.set_language(lang_setting)
                    # 加载文件夹设置
                    last_folder = config.get('last_folder')
                    if last_folder and Path(last_folder).exists():
                        self.root_folder = Path(last_folder)
                    # 加载Discogs Token
                    discogs_token = config.get('discogs_token')
                    if discogs_token:
                        self.DISCOGS_TOKEN = discogs_token
        except Exception as e:
            print(f"加载配置文件失败: {e}")
    
    def save_config(self):
        """保存当前选择的文件夹和语言设置到配置文件"""
        try:
            config = {}
            if self.config_file.exists():
                try:
                    with open(self.config_file, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                except:
                    pass
            
            # 保存语言设置
            config['language'] = self.lang.current_lang
            
            # 保存Discogs Token
            if self.DISCOGS_TOKEN:
                config['discogs_token'] = self.DISCOGS_TOKEN
            
            # 保存文件夹设置
            if self.root_folder:
                config['last_folder'] = str(self.root_folder)
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存配置文件失败: {e}")
    
    def update_tree_headers(self):
        """更新表格表头文本"""
        column_translations = {
            'folder_name': self.lang.t('folder_name'),
            'artist': self.lang.t('artist'),
            'album_name': self.lang.t('album_name'),
            'year': self.lang.t('year'),
            'status': self.lang.t('status'),
            'suggested_name': self.lang.t('suggested_name'),
        }
        
        for col in self.tree_columns:
            self.tree.heading(col, text=column_translations.get(col, col))
    
    def update_context_menu(self):
        """更新右键菜单文本"""
        # 清空现有菜单项
        self.context_menu.items = []
        
        # 重新添加菜单项（使用翻译）
        self.context_menu.add_command(self.lang.t('view_details'), self.view_details)
        self.context_menu.add_command(self.lang.t('single_search'), self.single_search)
        self.context_menu.add_command(self.lang.t('manual_input'), self.manual_input)
        self.context_menu.add_command(self.lang.t('rename_folder'), self.rename_folder)
        self.context_menu.add_separator()
        self.context_menu.add_command(self.lang.t('open_folder'), self.open_folder)
        self.context_menu.add_command(self.lang.t('play_folder'), self.play_folder)
    
    def on_language_changed(self, event=None):
        """语言切换事件处理"""
        # 获取选中的显示名称
        selected_display_name = self.lang_var.get()
        # 找到对应的语言代码
        if selected_display_name in self.lang_combo['values']:
            selected_index = list(self.lang_combo['values']).index(selected_display_name)
            if 0 <= selected_index < len(self.lang_values):
                selected_lang_code = self.lang_values[selected_index]
                self.lang.set_language(selected_lang_code)
                self.save_config()
                
                # 重新设置窗口标题
                self.root.title(self.lang.t('app_title') + ' V3.3')
                
                # 更新Token状态标签
                token_status = self.lang.t('token_configured') if (self.DISCOGS_TOKEN and self.DISCOGS_TOKEN != "YOUR_DISCOGS_TOKEN_HERE") else self.lang.t('token_not_configured')
                token_color = self.success_color if token_status == self.lang.t('token_configured') else self.error_color
                self.token_label_ref.config(text=f"{self.lang.t('token_status')}: {token_status}", foreground=token_color)
                
                # 更新按钮文本
                self.select_btn_ref.config(text=self.lang.t('select_folder'))
                self.start_btn_ref.config(text=self.lang.t('start_processing'))
                self.batch_rename_btn_ref.config(text=self.lang.t('batch_rename'))
                self.export_btn_ref.config(text=self.lang.t('export_excel'))
                
                # 更新语言标签
                self.lang_label.config(text=f"{self.lang.t('language')}:")
                
                # 更新表格表头
                self.update_tree_headers()
                
                # 更新右键菜单
                self.update_context_menu()
                
                # 更新状态栏
                if hasattr(self, 'status_var'):
                    current_status = self.status_var.get()
                    # 如果状态栏显示的是"就绪"，更新为翻译后的文本
                    if current_status == '就绪' or current_status == 'Ready' or current_status == self.lang.t('ready'):
                        self.status_var.set(self.lang.t('ready'))
                
                # 更新主列表中所有项的状态文本
                self.update_all_tree_items_status()
    
    def select_folder(self):
        """选择根文件夹"""
        folder = filedialog.askdirectory(title="选择包含音乐专辑文件夹的根目录")
        if folder:
            self.root_folder = Path(folder)
            self.save_config()  # 保存配置
            self.scan_folders()
    
    def scan_folders(self):
        """扫描文件夹下的所有子文件夹"""
        if not self.root_folder:
            return
        
        self.album_folders = []
        self.tree.delete(*self.tree.get_children())
        
        loaded_count = 0
        
        try:
            # 直接扫描选择文件夹下的所有子文件夹
            for idx, item in enumerate(self.root_folder.iterdir()):
                if item.is_dir():
                    folder_name = item.name
                    album_info = None
                    status_code = 'pending'
                    
                    # 检查是否有已处理的JSON文件
                    json_path = item / "album_info.json"
                    if json_path.exists():
                        try:
                            with open(json_path, 'r', encoding='utf-8') as f:
                                json_data = json.load(f)
                                album_info = AlbumInfo.from_dict(json_data)
                                status_code = 'completed'
                                loaded_count += 1
                        except Exception as e:
                            print(f"加载JSON文件失败 {json_path}: {e}")
                            # 如果加载失败，继续作为待处理
                    
                    status = self.get_status_text(status_code)
                    
                    self.album_folders.append((item, folder_name, album_info))
                    
                    # 根据是否有专辑信息显示不同的值
                    # 确定tag（状态标签）
                    if status_code == 'completed':
                        tag = 'completed'
                    elif status_code == 'searching':
                        tag = 'searching'
                    elif status_code == 'not_found':
                        tag = 'notfound'
                    else:
                        tag = 'pending'
                    
                    if album_info:
                        item = self.tree.insert('', tk.END, values=(
                            folder_name,
                            album_info.artist,
                            album_info.album_name,
                            album_info.year,
                            status,
                            album_info.get_suggested_folder_name()
                        ), tags=(tag,))
                    else:
                        item = self.tree.insert('', tk.END, values=(
                            folder_name, '', '', '', status, ''
                        ), tags=(tag,))
            
            status_msg = f"找到 {len(self.album_folders)} 个文件夹"
            if loaded_count > 0:
                status_msg += f"，已加载 {loaded_count} 个已处理记录"
            self.status_var.set(status_msg)
        except Exception as e:
            messagebox.showerror("错误", f"扫描文件夹时出错: {e}")
            self.status_var.set("扫描失败")
    
    def start_processing(self):
        """开始处理"""
        # 检查Token是否配置
        if not self.discogs_api:
            if not self.DISCOGS_TOKEN or self.DISCOGS_TOKEN == "YOUR_DISCOGS_TOKEN_HERE":
                messagebox.showwarning(self.lang.t('warning'), self.lang.t('configure_token') + "\n" + self.lang.t('configure_token_in_config'))
                return
            self.discogs_api = DiscogsAPI(self.DISCOGS_TOKEN)
        
        if not self.album_folders:
            messagebox.showwarning("警告", "请先选择文件夹")
            return
        
        # 在新线程中处理
        self.processing_thread = threading.Thread(target=self.process_folders, daemon=True)
        self.processing_thread.start()
    
    def process_folders(self):
        """处理所有文件夹"""
        total = len(self.album_folders)
        # 统计待处理的文件夹数量
        pending_count = sum(1 for _, _, info in self.album_folders if not info)
        processed_count = 0
        
        for idx, (folder_path, folder_name, album_info) in enumerate(self.album_folders):
            # 如果已经有专辑信息（已完成状态），跳过
            if album_info:
                continue
            
            processed_count += 1
            
            # 更新进度（只计算待处理的文件夹）
            progress = (processed_count / pending_count * 100) if pending_count > 0 else 100
            self.root.after(0, lambda p=progress: self.progress_var.set(p))
            self.root.after(0, lambda i=idx, pc=processed_count, tc=pending_count: 
                self.update_status(self.lang.t('processing', current=pc, total=tc) + f": {self.album_folders[i][1]}"))
            
            # 更新状态为搜索中
            self.root.after(0, lambda i=idx: self.update_tree_item(i, status='searching'))
            
            # 搜索Discogs
            results = self.discogs_api.search(folder_name)
            
            # 如果只有一个结果，自动选择
            if len(results) == 1:
                album_info = self.process_release(results[0], folder_path)
                if album_info:
                    self.album_folders[idx] = (folder_path, folder_name, album_info)
                    suggested_name = album_info.get_suggested_folder_name()
                    self.root.after(0, lambda i=idx, s=suggested_name, a=album_info: 
                        self.update_tree_item(i, status='completed', album_info=a, suggested=s))
            else:
                # 多个结果或未找到，显示选择对话框 - 暂停处理，等待用户选择
                # 确保没有其他对话框正在显示
                while self.selection_dialog_active:
                    time.sleep(0.1)
                
                # 设置对话框激活标志
                self.selection_dialog_active = True
                self.selection_result = None  # 重置选择结果
                self.waiting_for_selection.clear()  # 清除事件
                
                # 在主线程中显示对话框（使用after_idle确保立即执行）
                self.root.after_idle(lambda i=idx, r=results, q=folder_name: self.show_selection_dialog(i, r, q))
                
                # 等待用户选择（最多等待5分钟，避免无限等待）
                self.waiting_for_selection.wait(timeout=300)
                
                # 获取用户选择的结果
                if self.selection_result:
                    album_info = self.process_release(self.selection_result, folder_path)
                    if album_info:
                        self.album_folders[idx] = (folder_path, folder_name, album_info)
                        suggested_name = album_info.get_suggested_folder_name()
                        self.root.after(0, lambda i=idx, s=suggested_name, a=album_info: 
                            self.update_tree_item(i, status='completed', album_info=a, suggested=s))
                    self.selection_result = None
                else:
                    # 用户取消或未选择，更新状态
                    if not results:
                        self.root.after(0, lambda i=idx: self.update_tree_item(i, status='not_found'))
                    else:
                        # 有结果但用户取消，保持pending状态
                        self.root.after(0, lambda i=idx: self.update_tree_item(i, status='pending'))
                
                # 清除对话框激活标志
                self.selection_dialog_active = False
            
            # 避免API速率限制
            time.sleep(1.2)
        
        self.root.after(0, lambda: self.status_var.set("处理完成"))
        self.root.after(0, lambda: self.progress_var.set(0))
    
    def process_release(self, release_data: Dict, folder_path: Path) -> Optional[AlbumInfo]:
        """处理单个专辑信息"""
        album_info = AlbumInfo(release_data)
        
        # 获取详细信息
        if album_info.release_id:
            details = self.discogs_api.get_release_details(album_info.release_id)
            if details:
                album_info.details = details
                album_info.notes = details.get('notes', '')
                
                # 提取曲目表
                tracklist = details.get('tracklist', [])
                album_info.tracklist = []
                for track in tracklist:
                    track_info = {
                        '位置': track.get('position', ''),
                        '标题': track.get('title', ''),
                        '时长': track.get('duration', '')
                    }
                    album_info.tracklist.append(track_info)
                
                # 获取所有图片
                images = details.get('images', [])
                album_info.images = [img.get('uri', '') for img in images if img.get('uri')]
        
        # 下载所有图片
        image_count = 0
        downloaded_urls = set()  # 跟踪已下载的URL，避免重复
        
        # 下载封面图片（如果存在且不在images列表中）
        if album_info.cover_image and album_info.cover_image not in album_info.images:
            cover_path = folder_path / "cover.jpg"
            if self.discogs_api.download_image(album_info.cover_image, cover_path):
                image_count += 1
                downloaded_urls.add(album_info.cover_image)
        
        # 下载所有图片（包括封面）
        if album_info.images:
            for idx, img_url in enumerate(album_info.images):
                if img_url and img_url not in downloaded_urls:
                    # 确定文件扩展名
                    ext = 'jpg'
                    if '.png' in img_url.lower():
                        ext = 'png'
                    elif '.gif' in img_url.lower():
                        ext = 'gif'
                    elif '.webp' in img_url.lower():
                        ext = 'webp'
                    
                    # 如果是第一张图片且没有单独的cover，保存为cover.jpg
                    if idx == 0 and not album_info.cover_image:
                        img_path = folder_path / "cover.jpg"
                    else:
                        img_path = folder_path / f"image_{idx+1}.{ext}"
                    
                    if self.discogs_api.download_image(img_url, img_path):
                        image_count += 1
                        downloaded_urls.add(img_url)
        
        # 如果只有cover_image，也下载它
        if album_info.cover_image and album_info.cover_image not in downloaded_urls:
            cover_path = folder_path / "cover.jpg"
            if self.discogs_api.download_image(album_info.cover_image, cover_path):
                image_count += 1
        
        # 保存文本信息
        info_path = folder_path / "album_info.json"
        with open(info_path, 'w', encoding='utf-8') as f:
            json.dump(album_info.to_dict(), f, ensure_ascii=False, indent=2)
        
        return album_info
    
    def show_selection_dialog(self, idx: int, results: List[Dict], search_query: str = None):
        """显示选择对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title(self.lang.t('select_album'))
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.root)
        dialog.grab_set()  # 模态对话框，禁止操作主窗体
        
        # 设置窗口大小并居中显示（相对于主窗体）
        dialog_width = 960
        dialog_height = 600
        
        # 先设置大小，再居中
        dialog.geometry(f"{dialog_width}x{dialog_height}")
        dialog.update_idletasks()
        self.center_dialog(dialog, dialog_width, dialog_height)
        
        # 保存对话框引用，以便跟随主窗体移动（在完全创建后添加）
        self.open_dialogs.append((dialog, dialog_width, dialog_height))
        
        # 对话框关闭时从列表中移除
        def on_dialog_close():
            try:
                self.open_dialogs.remove((dialog, dialog_width, dialog_height))
            except:
                pass
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
        
        folder_name = self.album_folders[idx][1]
        if search_query is None:
            search_query = folder_name
        
        # 文件夹名称标签（可点击，点击后填入搜索框）
        folder_label_frame = tk.Frame(dialog, bg=self.bg_color)
        folder_label_frame.pack(pady=5)
        
        folder_label_text = tk.Label(folder_label_frame, text=f"{self.lang.t('folder')}: ", 
                                     font=('Arial', 10, 'bold'),
                                     background=self.bg_color, foreground=self.text_color)
        folder_label_text.pack(side=tk.LEFT)
        
        folder_name_label = tk.Label(folder_label_frame, text=folder_name,
                                     font=('Arial', 10, 'bold'),
                                     background=self.bg_color, foreground=self.accent_color,
                                     cursor='hand2')
        folder_name_label.pack(side=tk.LEFT)
        
        # 点击文件夹名称时，将其填入搜索框
        def on_folder_name_click(event):
            search_var.set(folder_name)
            search_entry.focus_set()
            search_entry.select_range(0, tk.END)  # 选中所有文本，方便修改
        
        folder_name_label.bind('<Button-1>', on_folder_name_click)
        
        # 搜索关键词输入框和重查按钮
        search_frame = ttk.Frame(dialog)
        search_frame.pack(pady=10, padx=10, fill=tk.X)
        
        ttk.Label(search_frame, text=f"{self.lang.t('search_keyword')}:", 
                 background=self.bg_color, foreground=self.text_color).pack(side=tk.LEFT, padx=5)
        
        search_var = tk.StringVar(value=search_query)
        search_entry = tk.Entry(search_frame, textvariable=search_var, width=30,
                               bg=self.secondary_bg, fg=self.text_color,
                               insertbackground=self.text_color,
                               borderwidth=0, highlightthickness=0)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 存储当前结果列表（用于更新）
        current_results = results.copy()
        listbox_ref = [None]  # 使用列表来存储引用
        
        def refresh_search():
            """重新搜索"""
            new_query = search_var.get().strip()
            if not new_query:
                return
            
            if not self.discogs_api:
                if not self.DISCOGS_TOKEN or self.DISCOGS_TOKEN == "YOUR_DISCOGS_TOKEN_HERE":
                    messagebox.showwarning(self.lang.t('warning'), self.lang.t('configure_token'))
                    return
                self.discogs_api = DiscogsAPI(self.DISCOGS_TOKEN)
            
            # 清空列表
            if listbox_ref[0]:
                listbox_ref[0].delete(0, tk.END)
            
            # 搜索
            new_results = self.discogs_api.search(new_query)
            current_results.clear()
            current_results.extend(new_results)
            
            # 更新列表
            if listbox_ref[0]:
                for result in new_results:
                    title = result.get('title', '')
                    year = result.get('year', '')
                    label = result.get('label', [])
                    label_str = ', '.join([l.get('name', '') if isinstance(l, dict) else str(l) for l in label[:2]]) if label else ''
                    display_text = f"{title} ({year}) - {label_str}"
                    listbox_ref[0].insert(tk.END, display_text)
                
                # 更新提示文本
                if new_results:
                    status_label.config(text=self.lang.t('found_results', count=len(new_results)) + ":")
                else:
                    status_label.config(text=self.lang.t('no_results') + ":")
        
        ttk.Button(search_frame, text=self.lang.t('re_search'), command=refresh_search).pack(side=tk.LEFT, padx=5)
        
        # 状态标签
        if results:
            status_text = self.lang.t('found_results', count=len(results)) + ":"
        else:
            status_text = self.lang.t('no_results') + ":"
        
        status_label = ttk.Label(dialog, text=status_text,
                                font=('Arial', 9),
                                background=self.bg_color, foreground=self.text_color)
        status_label.pack(pady=5)
        
        # 列表框
        listbox_frame = ttk.Frame(dialog)
        listbox_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        listbox = tk.Listbox(listbox_frame, height=18, font=('Arial', 9),
                            bg=self.secondary_bg, fg=self.text_color,
                            selectbackground=self.accent_color,
                            selectforeground='white',
                            borderwidth=0,
                            highlightthickness=0)
        scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=listbox.yview)
        listbox.configure(yscrollcommand=scrollbar.set)
        
        listbox_ref[0] = listbox  # 保存引用
        
        for result in results:
            title = result.get('title', '')
            year = result.get('year', '')
            label = result.get('label', [])
            label_str = ', '.join([l.get('name', '') if isinstance(l, dict) else str(l) for l in label[:2]]) if label else ''
            display_text = f"{title} ({year}) - {label_str}"
            listbox.insert(tk.END, display_text)
        
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        def on_select():
            selection = listbox.curselection()
            if selection and current_results:
                self.selection_result = current_results[selection[0]]
                try:
                    self.open_dialogs.remove((dialog, dialog_width, dialog_height))
                except:
                    pass
                self.selection_dialog_active = False  # 清除对话框激活标志
                dialog.destroy()
                self.waiting_for_selection.set()  # 通知处理线程继续
        
        def on_double_click(event):
            """双击选择"""
            selection = listbox.curselection()
            if selection and current_results:
                on_select()
        
        def on_cancel():
            self.selection_result = None
            try:
                self.open_dialogs.remove((dialog, dialog_width, dialog_height))
            except:
                pass
            self.selection_dialog_active = False  # 清除对话框激活标志
            dialog.destroy()
            self.waiting_for_selection.set()  # 通知处理线程继续（取消）
        
        # 绑定双击事件和回车键
        listbox.bind('<Double-Button-1>', on_double_click)
        search_entry.bind('<Return>', lambda e: refresh_search())
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text=self.lang.t('confirm'), command=on_select).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.lang.t('cancel'), command=on_cancel).pack(side=tk.LEFT, padx=5)
        
        # 注意：dialog.wait_window() 会阻塞主线程，但处理线程在等待 waiting_for_selection
        # 所以这里不需要 wait_window，因为处理线程会等待事件
        # dialog.wait_window()  # 已移除，由处理线程的 waiting_for_selection.wait() 控制
    
    def get_status_text(self, status_code):
        """获取状态文本的翻译"""
        status_map = {
            'pending': self.lang.t('pending'),
            'searching': self.lang.t('searching'),
            'completed': self.lang.t('completed'),
            'not_found': self.lang.t('not_found'),
        }
        return status_map.get(status_code, status_code)
    
    def update_tree_item(self, idx: int, status: str = None, album_info: AlbumInfo = None, suggested: str = None):
        """更新树形视图项"""
        folder_path, folder_name, current_info = self.album_folders[idx]
        
        if album_info:
            current_info = album_info
        
        # 确定状态和tag
        if status:
            # 统一处理状态码（支持状态码和翻译文本）
            if status in ['completed', self.lang.t('completed'), '已完成']:
                status_code = 'completed'
            elif status in ['searching', self.lang.t('searching'), '搜索中']:
                status_code = 'searching'
            elif status in ['not_found', self.lang.t('not_found'), '未找到']:
                status_code = 'not_found'
            elif status in ['pending', self.lang.t('pending'), '待处理']:
                status_code = 'pending'
            else:
                # 默认根据是否有专辑信息判断
                status_code = 'completed' if current_info else 'pending'
        else:
            status_code = 'completed' if current_info else 'pending'
        
        final_status = self.get_status_text(status_code)
        
        if status_code == 'completed':
            tag = 'completed'
        elif status_code == 'searching':
            tag = 'searching'
        elif status_code == 'not_found':
            tag = 'notfound'
        else:
            tag = 'pending'
        
        if current_info:
            values = (
                folder_name,
                current_info.artist,
                current_info.album_name,
                current_info.year,
                final_status,
                suggested or current_info.get_suggested_folder_name()
            )
        else:
            values = (
                folder_name,
                '',
                '',
                '',
                final_status,
                ''
            )
        
        # 通过索引直接获取对应的树视图项
        children = list(self.tree.get_children())
        if 0 <= idx < len(children):
            item = children[idx]
            self.tree.item(item, values=values, tags=(tag,))
        else:
            # 如果索引不匹配，回退到名称匹配方式
            for item in self.tree.get_children():
                if self.tree.item(item, 'values')[0] == folder_name:
                    self.tree.item(item, values=values, tags=(tag,))
                    break
    
    def update_all_tree_items_status(self):
        """更新主列表中所有项的状态文本（用于语言切换时）"""
        for idx in range(len(self.album_folders)):
            folder_path, folder_name, album_info = self.album_folders[idx]
            
            # 检查当前显示的状态来确定状态码
            children = list(self.tree.get_children())
            if 0 <= idx < len(children):
                item = children[idx]
                current_values = self.tree.item(item, 'values')
                current_tags = self.tree.item(item, 'tags')
                
                # 根据tag确定状态码（tag更可靠）
                if current_tags:
                    tag = current_tags[0]
                    if tag == 'completed':
                        status_code = 'completed'
                    elif tag == 'searching':
                        status_code = 'searching'
                    elif tag == 'notfound':
                        status_code = 'not_found'
                    else:
                        status_code = 'pending'
                else:
                    # 如果没有tag，根据是否有专辑信息判断
                    status_code = 'completed' if album_info else 'pending'
            else:
                # 如果索引不匹配，根据是否有专辑信息判断
                status_code = 'completed' if album_info else 'pending'
            
            # 重新更新该项，使用状态码
            self.update_tree_item(idx, status=status_code, album_info=album_info)
    
    def update_status(self, message: str):
        """更新状态栏"""
        self.status_var.set(message)
    
    def show_context_menu(self, event):
        """显示右键菜单"""
        item = self.tree.selection()[0] if self.tree.selection() else None
        if item:
            self.context_menu.post(event.x_root, event.y_root)
    
    def on_double_click(self, event):
        """双击事件"""
        self.view_details()
    
    def view_details(self):
        """查看详情"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.tree.item(item, 'values')
        folder_name = values[0]
        
        # 找到对应的专辑信息和文件夹路径
        for folder_path, name, album_info in self.album_folders:
            if name == folder_name:
                if album_info:
                    self.show_details_dialog(album_info, folder_path)
                else:
                    self.show_toast(self.lang.t('not_processed'), duration=2000)
                break
    
    def show_details_dialog(self, album_info: AlbumInfo, folder_path: Path):
        """显示详情对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title(self.lang.t('details_title'))
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.root)
        dialog.grab_set()  # 模态对话框，禁止操作主窗体
        
        # 设置窗口大小并居中显示（相对于主窗体）
        dialog_width = 600
        dialog_height = 700
        
        # 先设置大小，再居中
        dialog.geometry(f"{dialog_width}x{dialog_height}")
        dialog.update_idletasks()
        self.center_dialog(dialog, dialog_width, dialog_height)
        
        # 保存对话框引用，以便跟随主窗体移动（在完全创建后添加）
        self.open_dialogs.append((dialog, dialog_width, dialog_height))
        
        # 对话框关闭时从列表中移除
        def on_dialog_close():
            try:
                self.open_dialogs.remove((dialog, dialog_width, dialog_height))
            except:
                pass
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
        
        # 检查文件夹中是否有封面图片
        cover_image_path = None
        cover_extensions = ['cover.jpg', 'cover.png', 'cover.jpeg', 'cover.gif', 'cover.webp']
        for ext in cover_extensions:
            test_path = folder_path / ext
            if test_path.exists():
                cover_image_path = test_path
                break
        
        # 如果没有找到cover，尝试找image_1等
        if not cover_image_path:
            for ext in ['jpg', 'png', 'jpeg', 'gif', 'webp']:
                test_path = folder_path / f"image_1.{ext}"
                if test_path.exists():
                    cover_image_path = test_path
                    break
        
        # 如果有封面图片，在顶部显示
        if cover_image_path:
            try:
                # 使用PIL加载图片并调整大小
                img = Image.open(cover_image_path)
                # 限制最大尺寸为300x300
                img.thumbnail((300, 300), Image.Resampling.LANCZOS)
                
                # 转换为tkinter可用的格式
                photo = ImageTk.PhotoImage(img)
                
                # 创建图片标签
                image_frame = ttk.Frame(dialog)
                image_frame.pack(pady=10)
                
                image_label = ttk.Label(image_frame, image=photo)
                image_label.image = photo  # 保持引用
                image_label.pack()
            except Exception as e:
                print(f"加载封面图片失败: {e}")
        
        text_widget = tk.Text(dialog, wrap=tk.WORD, padx=10, pady=10,
                             bg=self.secondary_bg, fg=self.text_color,
                             insertbackground=self.text_color,
                             borderwidth=0,
                             highlightthickness=0)
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        # 格式化曲目表
        tracklist_text = ""
        if album_info.tracklist:
            tracklist_text = f"\n\n{self.lang.t('tracklist_label')}:\n"
            for track in album_info.tracklist:
                position = track.get('位置', '')
                title = track.get('标题', '')
                duration = track.get('时长', '')
                tracklist_text += f"  {position}. {title}"
                if duration:
                    tracklist_text += f" ({duration})"
                tracklist_text += "\n"
        
        details_text = f"""
{self.lang.t('artist_label')}: {album_info.artist}
{self.lang.t('album_name_label')}: {album_info.album_name}
{self.lang.t('year_label')}: {album_info.year}
{self.lang.t('label_label')}: {', '.join(album_info.label_names)}
{self.lang.t('catalog_no_label')}: {album_info.catalog_no}
{self.lang.t('genre_label')}: {', '.join(album_info.genre) if isinstance(album_info.genre, list) else album_info.genre}
{self.lang.t('style_label')}: {', '.join(album_info.style) if isinstance(album_info.style, list) else album_info.style}
{self.lang.t('country_label')}: {album_info.country}
{self.lang.t('discogs_id_label')}: {album_info.release_id}
{tracklist_text}
{self.lang.t('notes_label')}:
{album_info.notes if album_info.notes else self.lang.t('no_results')}
        """
        
        text_widget.insert('1.0', details_text.strip())
        text_widget.config(state=tk.DISABLED)
        
        # 按钮区域
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        def copy_to_clipboard():
            """复制信息到剪贴板"""
            dialog.clipboard_clear()
            dialog.clipboard_append(details_text.strip())
            dialog.update()  # 确保剪贴板更新
            self.show_toast(self.lang.t('info_copied'), duration=1500)
        
        def on_close():
            try:
                self.open_dialogs.remove((dialog, dialog_width, dialog_height))
            except:
                pass
            dialog.destroy()
        
        ttk.Button(button_frame, text=self.lang.t('copy_info'), command=copy_to_clipboard).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.lang.t('close'), command=on_close).pack(side=tk.LEFT, padx=5)
    
    def single_search(self):
        """单次查询 - 仅对该条信息检索discogs（支持重新查询已处理的文件夹）"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.tree.item(item, 'values')
        folder_name = values[0]
        
        # 找到对应的文件夹（注意：即使已有album_info也可以重新查询）
        for idx, (folder_path, name, existing_info) in enumerate(self.album_folders):
            if name == folder_name:
                if not self.discogs_api:
                    if not self.DISCOGS_TOKEN or self.DISCOGS_TOKEN == "YOUR_DISCOGS_TOKEN_HERE":
                        messagebox.showwarning(self.lang.t('warning'), self.lang.t('configure_token') + "\n" + self.lang.t('configure_token_in_config'))
                        return
                    self.discogs_api = DiscogsAPI(self.DISCOGS_TOKEN)
                
                # 更新状态为搜索中
                self.update_tree_item(idx, status='searching')
                
                # 在后台线程中执行查询，避免阻塞UI
                def do_search():
                    try:
                        # 搜索Discogs（即使已有数据也可以重新查询）
                        results = self.discogs_api.search(folder_name)
                        
                        # 在主线程中处理结果
                        self.root.after(0, lambda r=results: self._handle_single_search_results(idx, folder_path, folder_name, existing_info, r))
                    except Exception as e:
                        # 处理错误
                        self.root.after(0, lambda: self._handle_single_search_error(idx, existing_info, e))
                
                # 启动后台线程
                search_thread = threading.Thread(target=do_search, daemon=True)
                search_thread.start()
                break
    
    def _handle_single_search_results(self, idx: int, folder_path: Path, folder_name: str, existing_info, results):
        """处理单个查询的结果（在主线程中调用）"""
        # 如果只有一个结果，自动选择
        if len(results) == 1:
            # 在后台线程中处理release，避免阻塞UI
            def process_release():
                try:
                    album_info = self.process_release(results[0], folder_path)
                    if album_info:
                        # 在主线程中更新UI
                        self.root.after(0, lambda a=album_info: self._update_single_search_success(idx, folder_path, folder_name, a))
                except Exception as e:
                    self.root.after(0, lambda: self._handle_single_search_error(idx, existing_info, e))
            
            process_thread = threading.Thread(target=process_release, daemon=True)
            process_thread.start()
        else:
            # 多个结果或未找到，显示选择对话框
            # 确保没有其他对话框正在显示
            if self.selection_dialog_active:
                # 如果对话框正在显示，等待它关闭
                self.root.after(100, lambda r=results: self._handle_single_search_results(idx, folder_path, folder_name, existing_info, r))
                return
            
            # 设置对话框激活标志
            self.selection_dialog_active = True
            self.selection_result = None  # 重置选择结果
            self.waiting_for_selection.clear()  # 清除事件
            
            # 显示选择对话框
            self.show_selection_dialog(idx, results, folder_name)
            
            # 在后台线程中等待用户选择
            def wait_for_selection():
                # 等待用户选择（最多等待5分钟）
                self.waiting_for_selection.wait(timeout=300)
                
                # 在主线程中处理选择结果
                self.root.after(0, lambda: self._handle_single_search_selection(idx, folder_path, folder_name, existing_info, results))
            
            wait_thread = threading.Thread(target=wait_for_selection, daemon=True)
            wait_thread.start()
    
    def _update_single_search_success(self, idx: int, folder_path: Path, folder_name: str, album_info):
        """更新单个查询成功的结果（在主线程中调用）"""
        # 更新数据（覆盖原有的album_info）
        self.album_folders[idx] = (folder_path, folder_name, album_info)
        suggested_name = album_info.get_suggested_folder_name()
        # 使用状态码更新，确保多语言支持
        self.update_tree_item(idx, status='completed', album_info=album_info, suggested=suggested_name)
        self.show_toast(self.lang.t('search_success'), duration=1500)
    
    def _handle_single_search_selection(self, idx: int, folder_path: Path, folder_name: str, existing_info, results):
        """处理单个查询的用户选择结果（在主线程中调用）"""
        # 检查用户是否选择了结果
        if self.selection_result:
            # 在后台线程中处理release
            def process_selected():
                try:
                    album_info = self.process_release(self.selection_result, folder_path)
                    if album_info:
                        # 在主线程中更新UI
                        self.root.after(0, lambda a=album_info: self._update_single_search_success(idx, folder_path, folder_name, a))
                    self.selection_result = None
                    self.selection_dialog_active = False
                except Exception as e:
                    self.root.after(0, lambda: self._handle_single_search_error(idx, existing_info, e))
                    self.selection_dialog_active = False
            
            process_thread = threading.Thread(target=process_selected, daemon=True)
            process_thread.start()
        elif not results:
            # 未找到结果
            self.update_tree_item(idx, status='not_found')
            self.selection_dialog_active = False
        else:
            # 有结果但用户取消，恢复原状态
            if existing_info:
                # 如果有原有数据，恢复为已完成状态
                self.update_tree_item(idx, status='completed', album_info=existing_info)
            else:
                # 如果没有原有数据，恢复为待处理状态
                self.update_tree_item(idx, status='pending')
            self.selection_dialog_active = False
    
    def _handle_single_search_error(self, idx: int, existing_info, error):
        """处理单个查询的错误（在主线程中调用）"""
        # 恢复原状态
        if existing_info:
            self.update_tree_item(idx, status='completed', album_info=existing_info)
        else:
            self.update_tree_item(idx, status='pending')
        messagebox.showerror(self.lang.t('error'), f"{self.lang.t('search_failed')}: {error}")
    
    def manual_input(self):
        """手动录入 - 不查discogs，弹出录入界面"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.tree.item(item, 'values')
        folder_name = values[0]
        
        # 找到对应的文件夹
        for idx, (folder_path, name, _) in enumerate(self.album_folders):
            if name == folder_name:
                # 创建输入对话框
                dialog = tk.Toplevel(self.root)
                dialog.title(self.lang.t('manual_input_title'))
                dialog.configure(bg=self.bg_color)
                dialog.transient(self.root)
                dialog.grab_set()  # 模态对话框，禁止操作主窗体
                
                # 居中显示（相对于主窗体）
                dialog_width = 720
                dialog_height = 650
                
                # 先设置大小，再居中
                dialog.geometry(f"{dialog_width}x{dialog_height}")
                dialog.update_idletasks()
                self.center_dialog(dialog, dialog_width, dialog_height)
                
                # 保存对话框引用，以便跟随主窗体移动（在完全创建后添加）
                self.open_dialogs.append((dialog, dialog_width, dialog_height))
                
                # 对话框关闭时从列表中移除
                def on_dialog_close():
                    try:
                        self.open_dialogs.remove((dialog, dialog_width, dialog_height))
                    except:
                        pass
                    dialog.destroy()
                
                dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
                
                # 主容器 - 直接使用Frame，不使用滚动条
                scrollable_frame = tk.Frame(dialog, bg=self.bg_color)
                scrollable_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=3)
                
                fields = {}
                current_row = 0
                
                # ========== 第一部分：基本信息 ==========
                basic_frame = tk.Frame(scrollable_frame, bg=self.secondary_bg, relief=tk.FLAT)
                basic_frame.grid(row=current_row, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 3))
                current_row += 1
                
                tk.Label(basic_frame, text=self.lang.t('basic_info'), 
                        bg=self.secondary_bg, fg=self.accent_color,
                        font=('Arial', 10, 'bold')).pack(anchor='w', padx=12, pady=(3, 3))
                
                basic_fields = [
                    ('artist_label', 'artist'),
                    ('album_name_label', 'album_name'),
                    ('year_label', 'year')
                ]
                
                for label_key, field_key in basic_fields:
                    field_frame = tk.Frame(basic_frame, bg=self.secondary_bg)
                    field_frame.pack(fill=tk.X, padx=12, pady=1)
                    
                    tk.Label(field_frame, text=f"{self.lang.t(label_key)}:", 
                            bg=self.secondary_bg, fg=self.text_color,
                            font=('Arial', 9), width=12, anchor='w').pack(side=tk.LEFT)
                    
                    entry = tk.Entry(field_frame, width=90,
                                    bg=self.bg_color, fg=self.text_color,
                                    insertbackground=self.text_color,
                                    borderwidth=1, highlightthickness=1,
                                    highlightbackground=self.accent_bg,
                                    highlightcolor=self.accent_color,
                                    relief=tk.FLAT)
                    entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
                    fields[field_key] = entry
                
                # ========== 第二部分：发行信息 ==========
                release_frame = tk.Frame(scrollable_frame, bg=self.secondary_bg, relief=tk.FLAT)
                release_frame.grid(row=current_row, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 3))
                current_row += 1
                
                tk.Label(release_frame, text=self.lang.t('release_info'), 
                        bg=self.secondary_bg, fg=self.accent_color,
                        font=('Arial', 10, 'bold')).pack(anchor='w', padx=12, pady=(3, 3))
                
                release_fields = [
                    ('label_label', 'label'),
                    ('catalog_no_label', 'catalog_no'),
                    ('country_label', 'country')
                ]
                
                for label_key, field_key in release_fields:
                    field_frame = tk.Frame(release_frame, bg=self.secondary_bg)
                    field_frame.pack(fill=tk.X, padx=12, pady=1)
                    
                    tk.Label(field_frame, text=f"{self.lang.t(label_key)}:", 
                            bg=self.secondary_bg, fg=self.text_color,
                            font=('Arial', 9), width=12, anchor='w').pack(side=tk.LEFT)
                    
                    entry = tk.Entry(field_frame, width=90,
                                    bg=self.bg_color, fg=self.text_color,
                                    insertbackground=self.text_color,
                                    borderwidth=1, highlightthickness=1,
                                    highlightbackground=self.accent_bg,
                                    highlightcolor=self.accent_color,
                                    relief=tk.FLAT)
                    entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
                    fields[field_key] = entry
                
                # ========== 第三部分：分类信息 ==========
                category_frame = tk.Frame(scrollable_frame, bg=self.secondary_bg, relief=tk.FLAT)
                category_frame.grid(row=current_row, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 3))
                current_row += 1
                
                tk.Label(category_frame, text=self.lang.t('category_info'), 
                        bg=self.secondary_bg, fg=self.accent_color,
                        font=('Arial', 10, 'bold')).pack(anchor='w', padx=12, pady=(3, 3))
                
                category_fields = [
                    ('genre_label', 'genre'),
                    ('style_label', 'style')
                ]
                
                for label_key, field_key in category_fields:
                    field_frame = tk.Frame(category_frame, bg=self.secondary_bg)
                    field_frame.pack(fill=tk.X, padx=12, pady=1)
                    
                    tk.Label(field_frame, text=f"{self.lang.t(label_key)}:", 
                            bg=self.secondary_bg, fg=self.text_color,
                            font=('Arial', 9), width=12, anchor='w').pack(side=tk.LEFT)
                    
                    entry = tk.Entry(field_frame, width=90,
                                    bg=self.bg_color, fg=self.text_color,
                                    insertbackground=self.text_color,
                                    borderwidth=1, highlightthickness=1,
                                    highlightbackground=self.accent_bg,
                                    highlightcolor=self.accent_color,
                                    relief=tk.FLAT)
                    entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
                    fields[field_key] = entry
                
                # ========== 第四部分：其他信息 ==========
                other_frame = tk.Frame(scrollable_frame, bg=self.secondary_bg, relief=tk.FLAT)
                other_frame.grid(row=current_row, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 3))
                current_row += 1
                
                tk.Label(other_frame, text=self.lang.t('other_info'), 
                        bg=self.secondary_bg, fg=self.accent_color,
                        font=('Arial', 10, 'bold')).pack(anchor='w', padx=12, pady=(3, 3))
                
                # Discogs ID
                discogs_frame = tk.Frame(other_frame, bg=self.secondary_bg)
                discogs_frame.pack(fill=tk.X, padx=12, pady=1)
                
                tk.Label(discogs_frame, text=f"{self.lang.t('discogs_id_label')}:", 
                        bg=self.secondary_bg, fg=self.text_color,
                        font=('Arial', 9), width=12, anchor='w').pack(side=tk.LEFT)
                
                discogs_entry = tk.Entry(discogs_frame, width=90,
                                        bg=self.bg_color, fg=self.text_color,
                                        insertbackground=self.text_color,
                                        borderwidth=1, highlightthickness=1,
                                        highlightbackground=self.accent_bg,
                                        highlightcolor=self.accent_color,
                                        relief=tk.FLAT)
                discogs_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
                fields['discogs_id'] = discogs_entry
                
                # 备注信息
                notes_frame = tk.Frame(other_frame, bg=self.secondary_bg)
                notes_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=1)
                
                tk.Label(notes_frame, text=f"{self.lang.t('notes_label')}:", 
                        bg=self.secondary_bg, fg=self.text_color,
                        font=('Arial', 9), anchor='w').pack(anchor='w', pady=(0, 3))
                
                notes_text = tk.Text(notes_frame, height=4, width=90,
                                    bg=self.bg_color, fg=self.text_color,
                                    insertbackground=self.text_color,
                                    borderwidth=1, highlightthickness=1,
                                    highlightbackground=self.accent_bg,
                                    highlightcolor=self.accent_color,
                                    relief=tk.FLAT, wrap=tk.WORD)
                notes_text.pack(fill=tk.BOTH, expand=True)
                fields['notes'] = notes_text
                
                # ========== 第五部分：曲目表 ==========
                tracklist_frame = tk.Frame(scrollable_frame, bg=self.secondary_bg, relief=tk.FLAT)
                tracklist_frame.grid(row=current_row, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 3))
                current_row += 1
                
                tk.Label(tracklist_frame, text=self.lang.t('tracklist'), 
                        bg=self.secondary_bg, fg=self.accent_color,
                        font=('Arial', 10, 'bold')).pack(anchor='w', padx=12, pady=(3, 3))
                
                # 曲目表提示
                hint_label = tk.Label(tracklist_frame, 
                                     text=self.lang.t('tracklist_hint'),
                                     bg=self.secondary_bg, fg=self.text_color,
                                     font=('Arial', 8), anchor='w', justify=tk.LEFT)
                hint_label.pack(anchor='w', padx=12, pady=(0, 3))
                
                tracklist_text = tk.Text(tracklist_frame, height=8, width=90,
                                        bg=self.bg_color, fg=self.text_color,
                                        insertbackground=self.text_color,
                                        borderwidth=1, highlightthickness=1,
                                        highlightbackground=self.accent_bg,
                                        highlightcolor=self.accent_color,
                                        relief=tk.FLAT, wrap=tk.WORD,
                                        font=('Consolas', 9))
                tracklist_text.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 3))
                fields['tracklist'] = tracklist_text
                
                scrollable_frame.columnconfigure(0, weight=1)
                
                def save_manual_input():
                    """保存手动录入的信息"""
                    try:
                        # 创建临时的release_data结构
                        release_data = {
                            'id': int(fields['discogs_id'].get()) if fields['discogs_id'].get().strip() else None,
                            'title': f"{fields['artist'].get()} - {fields['album_name'].get()}",
                            'year': fields['year'].get(),
                            'catno': fields['catalog_no'].get(),
                            'country': fields['country'].get(),
                            'genre': [g.strip() for g in fields['genre'].get().split(',') if g.strip()],
                            'style': [s.strip() for s in fields['style'].get().split(',') if s.strip()],
                            'label': [{'name': l.strip()} for l in fields['label'].get().split(',') if l.strip()],
                            'cover_image': '',
                            'thumb': ''
                        }
                        
                        # 创建AlbumInfo对象
                        album_info = AlbumInfo(release_data)
                        album_info.notes = fields['notes'].get('1.0', tk.END).strip()
                        
                        # 解析曲目表
                        tracklist_lines = tracklist_text.get('1.0', tk.END).strip().split('\n')
                        album_info.tracklist = []
                        for line in tracklist_lines:
                            if line.strip():
                                # 简单解析：位置. 标题 时长
                                parts = line.strip().split('.', 1)
                                if len(parts) == 2:
                                    position = parts[0].strip()
                                    rest = parts[1].strip()
                                    # 尝试分离标题和时长
                                    if '(' in rest and ')' in rest:
                                        title = rest.split('(')[0].strip()
                                        duration = rest.split('(')[1].split(')')[0].strip()
                                    else:
                                        title = rest
                                        duration = ''
                                    album_info.tracklist.append({
                                        '位置': position,
                                        '标题': title,
                                        '时长': duration
                                    })
                        
                        # 保存JSON文件
                        info_path = folder_path / "album_info.json"
                        with open(info_path, 'w', encoding='utf-8') as f:
                            json.dump(album_info.to_dict(), f, ensure_ascii=False, indent=2)
                        
                        # 更新数据
                        self.album_folders[idx] = (folder_path, folder_name, album_info)
                        suggested_name = album_info.get_suggested_folder_name()
                        self.update_tree_item(idx, status='completed', album_info=album_info, suggested=suggested_name)
                        
                        self.show_toast(self.lang.t('manual_input_success'), duration=1500)
                        try:
                            self.open_dialogs.remove((dialog, dialog_width, dialog_height))
                        except:
                            pass
                        dialog.destroy()
                    except Exception as e:
                        messagebox.showerror(self.lang.t('error'), f"{self.lang.t('save_failed')}: {e}")
                
                def on_cancel():
                    try:
                        self.open_dialogs.remove((dialog, dialog_width, dialog_height))
                    except:
                        pass
                    dialog.destroy()
                
                # 按钮区域 - 固定在底部
                button_frame = tk.Frame(dialog, bg=self.bg_color)
                button_frame.pack(fill=tk.X, padx=15, pady=(0, 10))
                
                # 分隔线
                separator = tk.Frame(button_frame, bg=self.accent_bg, height=1)
                separator.pack(fill=tk.X, pady=(0, 10))
                
                # 按钮容器
                btn_container = ttk.Frame(button_frame)
                btn_container.pack()
                
                # 保存按钮 - 使用ttk.Button与其他面板保持一致
                save_btn = ttk.Button(btn_container, text=self.lang.t('save'),
                                     command=save_manual_input)
                save_btn.pack(side=tk.LEFT, padx=(0, 5))
                
                # 取消按钮 - 使用ttk.Button与其他面板保持一致
                cancel_btn = ttk.Button(btn_container, text=self.lang.t('cancel'),
                                        command=on_cancel)
                cancel_btn.pack(side=tk.LEFT, padx=5)
                
                break
    
    def rename_folder(self):
        """重命名文件夹"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.tree.item(item, 'values')
        folder_name = values[0]
        suggested_name = values[5]
        
        if not suggested_name:
            messagebox.showwarning("警告", "没有建议的名称")
            return
        
        # 找到对应的文件夹
        for idx, (folder_path, name, album_info) in enumerate(self.album_folders):
            if name == folder_name:
                # 确保名称已清理（双重保险）
                if album_info:
                    cleaned_name = album_info.get_suggested_folder_name()
                else:
                    # 如果没有album_info，手动清理
                    cleaned_name = re.sub(r'[<>:"/\\|?*]', '_', suggested_name).strip(' .')
                
                new_path = folder_path.parent / cleaned_name
                
                if new_path.exists():
                    messagebox.showwarning("警告", f"目标文件夹已存在: {cleaned_name}")
                    return
                
                try:
                    folder_path.rename(new_path)
                    # 更新album_folders中的路径和名称，保留专辑信息
                    self.album_folders[idx] = (new_path, cleaned_name, album_info)
                    
                    # 更新树视图中的显示
                    self.update_tree_item(idx, status='completed' if album_info else 'pending', 
                                         album_info=album_info, 
                                         suggested=cleaned_name if album_info else '')
                    
                    self.show_toast(self.lang.t('folder_renamed', name=cleaned_name), duration=2000)
                except Exception as e:
                    messagebox.showerror("错误", f"重命名失败: {e}")
                break
    
    def batch_rename(self):
        """批量重命名所有已完成的文件夹"""
        rename_count = 0
        skipped_count = 0
        error_count = 0
        
        # 使用索引遍历，以便更新列表
        for idx in range(len(self.album_folders)):
            folder_path, folder_name, album_info = self.album_folders[idx]
            
            if not album_info:
                continue
            
            suggested_name = album_info.get_suggested_folder_name()
            if not suggested_name or suggested_name == folder_name:
                skipped_count += 1
                continue
            
            new_path = folder_path.parent / suggested_name
            
            if new_path.exists():
                skipped_count += 1
                continue
            
            try:
                folder_path.rename(new_path)
                # 更新album_folders中的路径和名称，保留专辑信息
                self.album_folders[idx] = (new_path, suggested_name, album_info)
                
                # 更新树视图中的显示
                self.update_tree_item(idx, status='completed', album_info=album_info, suggested=suggested_name)
                
                rename_count += 1
            except Exception as e:
                error_count += 1
                print(f"重命名失败 {folder_name}: {e}")
        
        self.show_toast(self.lang.t('batch_rename_complete', success=rename_count, skipped=skipped_count, failed=error_count), duration=2500)
    
    def open_folder(self):
        """打开文件夹"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.tree.item(item, 'values')
        folder_name = values[0]
        
        for folder_path, name, _ in self.album_folders:
            if name == folder_name:
                try:
                    if sys.platform == 'win32':
                        os.startfile(folder_path)
                    elif sys.platform == 'darwin':
                        os.system(f'open "{folder_path}"')
                    else:
                        os.system(f'xdg-open "{folder_path}"')
                except Exception as e:
                    messagebox.showerror("错误", f"无法打开文件夹: {e}")
                break
    
    def play_folder(self):
        """使用foobar2000播放文件夹"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.tree.item(item, 'values')
        folder_name = values[0]
        
        for folder_path, name, _ in self.album_folders:
            if name == folder_name:
                try:
                    if sys.platform == 'win32':
                        # foobar2000可能的安装路径
                        foobar_paths = [
                            r"C:\Program Files\foobar2000\foobar2000.exe",
                            r"C:\Program Files (x86)\foobar2000\foobar2000.exe",
                            r"C:\Users\{}\AppData\Local\foobar2000\foobar2000.exe".format(os.getenv('USERNAME')),
                            "foobar2000.exe"  # 如果在PATH中
                        ]
                        
                        foobar_exe = None
                        for path in foobar_paths:
                            if path == "foobar2000.exe":
                                # 尝试直接调用（如果在PATH中）
                                try:
                                    cmd = f'foobar2000.exe "{folder_path}" /play'
                                    subprocess.Popen(shlex.split(cmd), shell=False,
                                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                                    self.show_toast(self.lang.t('playing_in_foobar'), duration=1500)
                                    return
                                except FileNotFoundError:
                                    pass
                                except Exception as e:
                                    print(f"调用播放器出错: {e}")
                            elif os.path.exists(path):
                                foobar_exe = path
                                break
                        
                        if foobar_exe:
                            # 使用 /immediate 和 /play 参数
                            cmd = f'"{foobar_exe}" "{folder_path}" /play'
                            subprocess.Popen(shlex.split(cmd), shell=False,
                                           creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                            self.show_toast(self.lang.t('playing_in_foobar'), duration=1500)
                        else:
                            # 尝试直接调用foobar2000（可能在PATH中）
                            try:
                                cmd = f'foobar2000.exe "{folder_path}" /play'
                                subprocess.Popen(shlex.split(cmd), shell=False,
                                               creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                                self.show_toast(self.lang.t('playing_in_foobar'), duration=1500)
                            except FileNotFoundError:
                                self.show_toast(self.lang.t('foobar_not_found'), duration=2000)
                            except Exception as e:
                                print(f"调用播放器出错: {e}")
                                self.show_toast(self.lang.t('play_error', error=str(e)), duration=2000)
                    else:
                        # Linux/Mac系统
                        self.show_toast(self.lang.t('system_not_supported'), duration=2000)
                except Exception as e:
                    self.show_toast(f"播放失败: {str(e)}", duration=2000)
                break
    
    def export_excel(self):
        """导出到Excel"""
        # 收集所有已完成的专辑信息
        completed_albums = []
        for folder_path, folder_name, album_info in self.album_folders:
            if album_info:
                data = album_info.to_dict()
                data['文件夹名'] = folder_name
                data['文件夹路径'] = str(folder_path)
                completed_albums.append(data)
        
        if not completed_albums:
            messagebox.showwarning("警告", "没有可导出的数据")
            return
        
        # 选择保存位置
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="保存Excel文件"
        )
        
        if not filename:
            return
        
        # 创建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "专辑信息"
        
        # 设置表头
        headers = ['文件夹名', '音乐人', '专辑名', '出版年份', '唱片厂牌', '厂牌编号', 
                  '音乐风格', '风格标签', '备注信息', 'Discogs ID', '国家', '文件夹路径']
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row_idx, album_data in enumerate(completed_albums, 2):
            for col_idx, header in enumerate(headers, 1):
                value = album_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # 调整列宽
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            if header in ['文件夹路径', '备注信息']:
                ws.column_dimensions[col_letter].width = 40
            elif header in ['音乐人', '专辑名', '唱片厂牌']:
                ws.column_dimensions[col_letter].width = 25
            else:
                ws.column_dimensions[col_letter].width = 15
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存文件
        try:
            wb.save(filename)
            self.show_toast(self.lang.t('excel_saved', filename=filename), duration=2000)
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")


def main():
    root = tk.Tk()
    app = DiscMatcherApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

